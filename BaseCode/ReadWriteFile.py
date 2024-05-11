#!/usr/local/bin/python3
# coding: utf-8

###
# ReadWriteFile
# ファイルの読書用スクリプト
# 
###

#--- モジュールのインポート ---#
import sys
import csv
from openpyxl import load_workbook, Workbook

#--- 変数宣言 ---#
curScript = sys.argv[0]


#=== .txtファイル処理ブロック ===#
#--- .txtファイルの読込 ---#
# inputFile:入力'.txtファイル'(.txtファイルまでのパス)
def ReadTxtFile(inputFile):
    #--- 変数宣言 ---#

    #--- .txtファイルの中身を取得 ---#
    with open(inputFile, 'r') as f:
        allData = f.read()
    return allData


#--- .txtファイルの書込 ---#
# data:出力データ
# outputFile:出力'.txtファイル'(.txtファイルまでのパス)
def WriteTxtFile(data, outputFile):
    #--- 変数宣言 ---#

    #--- .txtファイル出力 ---#
    with open(outputFile, 'w') as f:
        f.write(data)


#=== .csvファイル処理ブロック ===#
#--- .csvファイル読込 ---#
# inputFile:入力'.csvファイル'(.csvファイルまでのパス)
def ReadCsvFile(inputFile):
    #--- 変数宣言 ---#
    data = []

    #--- .csvファイルの中身をリストとして取得 ---#
    with open(inputFile, 'r', encoding='utf-8', newline='') as file:
        reader = csv.reader(file)
        for row in reader:
            data.append(row)
    return data


#--- .tsvファイル読込 ---#
# inputFile:入力'.tsvファイル'(.tsvファイルまでのパス)
def ReadTsvFile(inputFile):
    #--- 変数宣言 ---#
    data = []

    #--- .tsvファイルの中身をリストとして取得 ---#
    with open(inputFile, 'r', encoding='utf-8', newline='') as file:
        reader = csv.reader(file, delimiter='\t')
        for row in reader:
            data.append(row)
    return data


#--- .csv(tsv)ファイル出力 ---#
# data:出力データ
# outputFile:出力'.csvファイル'(.csvファイルまでのパス)
# デリミタはタブ(\t)
def WriteTsvFile(data, outputFile):
    #--- 変数宣言 ---#

    #--- tsvファイル出力 ---#
    with open(outputFile, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f, delimiter='\t')

        # 各業のデータをCSVファイルに書込
        for row in data:
            # 各行のデータをタブで分割してリスト化し、CSVファイルに書込
            writer.writerow(row.split('\t'))


#=== .xlsxファイル処理ブロック ===#
#--- ('\t'区切りの)dataをExcelに新規出力 ---#
# data:出力データ
# outputFile:出力'.xlsxファイル'(.xlsxファイルまでのパス)
def WriteTabSprDataToExcelFile(data, outputFile):
    #--- 変数宣言 ---#

    # 新規ワークブックの作成
    wb = Workbook()
    ws = wb.active

    # 各行のデータをExcelシートに追加
    for row in data:
        # タブで分割してリストに格納
        formattedData = row.split('\t')
        # Excelシートにデータ行を追加
        ws.append(formattedData)

    # Excelファイルを保存
    wb.save(outputFile)


#--- Excelファイルの読込 ---#
# inputFile:入力'*.xlsx'ファイル(*.xlsxまでのパス)
def ReadExcel(inputFile):
    #--- 変数宣言 ---#

    wb = load_workbook(filename=inputFile, read_only=True)
    ws = wb.active
    listData = []

    for row in ws.iter_rows(values_only=True):
        listData.append(row)

    return listData


#--- 辞書型リストデータをExcelファイルへ書込 ---#
# data:辞書型リスト
def WriteDicListToExcel(data, outputFile):
    wb = Workbook()
    ws = wb.active

    # ヘッダーを取得
    headers = list(data[0].keys())
    
    # ヘッダーを書き込む
    ws.append(headers)
    
    # データを書き込む
    for item in data:
        row = [item[header] for header in headers]
        ws.append(row)

    # Excelファイルを保存
    wb.save(outputFile)


#--- Excelファイルからキーワードを含む行の抽出 ---#
# このメソッドは将来的に変更修正かけたい
def ExtractDataFromExcel(inputFile, colIndex, keyword):
    #--- 変数宣言 ---#

    # Excelファイルを開く
    wb = load_workbook(inputFile)
    # アクティブなシートを取得する
    ws = wb.active

    # データを保持するリスト
    listData = []

    for row in ws.iter_rows(values_only=True):
        if row[colIndex] == keyword:
            listData.append(row)
    
    return listData


#--- リストデータをExcelファイルへ書込 ---#
def ListDataToExcel(data, outputFile, headerStr=''):
    #--- 変数宣言 ---#
    startVal = 2

    wb = Workbook()
    ws = wb.active

    if (headerStr != ''):
        ws.append(headerStr)
        startVal = 1
    
    for idx, row in enumerate(data, start=startVal): # データの書込開始行
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=idx, column=col_idx, value=value)

    # Excelファイルを保存
    wb.save(outputFile)


#=== 設定ファイル読込ブロック ===#
#--- 色情報の設定ファイル(csv)を読み込み、辞書型の変数に値をセットし返却するメソッド ---#
# <Parameters>
# inputFile (str)           : 入力'.csvファイル'(.csvまでのパス)
#
# <Returns>
# colorDict(dict)           : 色情報辞書型変数(色名称:色コード)
def ReadColorInfoFile(inputFile):
    #--- 変数宣言 ---#
    colorDict = {}

    with open(inputFile, 'r', newline='') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            colorName = row['# ColorName']
            colorCode = row['ColorCode']
            colorDict[colorName] = colorCode

    return colorDict

#--- main関数 ---#
def main():
    print("Youは本スクリプト：{}を直接実行しました".format(curScript))

if __name__ == "__main__":
    main()
