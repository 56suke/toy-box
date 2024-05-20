#!/usr/local/bin/python3
# coding: utf-8

###
# FileReadWriter.py
# ファイルの読書用スクリプト
# 
###

#--- モジュールのインポート ---#
import sys
import csv
from openpyxl import load_workbook, Workbook

#--- 変数宣言 ---#
curScript = sys.argv[0]


#=== .txtファイル処理ブロック ===#
#--- .txtファイルの読込 ---#
# inputFile:入力'.txtファイル'(.txtファイルまでのパス)
def ReadTxtFile(inputFile):
    #--- 変数宣言 ---#

    #--- .txtファイルの中身を取得 ---#
    with open(inputFile, 'r') as file:
        data = [line.rstrip('\n') for line in file.readlines()]
    return data


#--- .txtファイルのデータ結合 ---#
def MergeTxtFiles(inputFileList):
    #--- 変数宣言 ---#
    mergedData = []

    for filePath in inputFileList:
        data = ReadTxtFile(filePath)
        mergedData.extend(data)
    return mergedData


#--- .txtファイルの書込 ---#
def WriteTxtFile(dataList, outputFile):
    #--- 変数宣言 ---#
    # リストの要素を文字列に連結
    dataStr = '\n'.join(dataList)

    #--- .txtファイル出力 ---#
    with open(outputFile, 'w') as file:
        file.write(dataStr)


#=== .csvファイル処理ブロック ===#
#--- .csvファイル読込 ---#
# inputFile:入力'.csvファイル'(.csvファイルまでのパス)
def ReadCsvFile(inputFile):
    #--- 変数宣言 ---#
    data = []

    #--- .csvファイルの中身をリストとして取得 ---#
    with open(inputFile, 'r', encoding='utf-8', newline='') as file:
        reader = csv.reader(file)
        for row in reader:
            data.append(row)
    return data


#--- .tsvファイル読込 ---#
# inputFile:入力'.tsvファイル'(.tsvファイルまでのパス)
def ReadTsvFile(inputFile):
    #--- 変数宣言 ---#
    data = []

    #--- .tsvファイルの中身をリストとして取得 ---#
    with open(inputFile, 'r', encoding='utf-8', newline='') as file:
        reader = csv.reader(file, delimiter='\t')
        for row in reader:
            data.append(row)
    return data


#--- .csv(tsv)ファイル出力 ---#
# data:出力データ
# outputFile:出力'.csvファイル'(.csvファイルまでのパス)
# デリミタはタブ(\t)
def WriteTsvFile(data, outputFile):
    #--- 変数宣言 ---#

    #--- tsvファイル出力 ---#
    with open(outputFile, 'w', encoding='utf-8', newline='') as file:
        writer = csv.writer(file, delimiter='\t')

        # 各業のデータをCSVファイルに書込
        for row in data:
            # 各行のデータをタブで分割してリスト化し、CSVファイルに書込
            writer.writerow(row.split('\t'))


#=== .xlsxファイル処理ブロック ===#
#--- tsvファイルからExcelファイルへの変換 ---#
def ConvertTsvToExcel(inputFile, outputFile):
    #--- 変数宣言 ---#

    # 新規ワークブックの作成
    wb = Workbook()
    ws = wb.active

    # .tsvファイルを開いて読み込む
    with open(inputFile, 'r', encoding='utf-8') as file:
        tsvReader = csv.reader(file, delimiter='\t')
        for row in tsvReader:
            ws.append(row) # 読み込んだ行をExcelシートに追加
    
    # Excelファイルを保存
    wb.save(outputFile)


#--- ('\t'区切りの)dataをExcelに新規出力 ---#
# data:出力データ
# outputFile:出力'.xlsxファイル'(.xlsxファイルまでのパス)
def WriteTabSprDataToExcelFile(data, outputFile):
    #--- 変数宣言 ---#

    # 新規ワークブックの作成
    wb = Workbook()
    ws = wb.active

    # 各行のデータをExcelシートに追加
    for row in data:
        # タブで分割してリストに格納
        formattedData = row.split('\t')
        # Excelシートにデータ行を追加
        ws.append(formattedData)

    # Excelファイルを保存
    wb.save(outputFile)


#--- Excelファイルの読込 ---#
# inputFile:入力'*.xlsx'ファイル(*.xlsxまでのパス)
def ReadExcel(inputFile):
    #--- 変数宣言 ---#

    wb = load_workbook(filename=inputFile, read_only=True)
    ws = wb.active
    listData = []

    for row in ws.iter_rows(values_only=True):
        listData.append(row)

    return listData


#--- 各行のIndex値も込みでExcelファイル読込 ---#
# minRos, startIndex共にヘッダを考慮し、default値を2とする
def ReadExcelWithIndex(inputFile, minRow=2, startIndex=2):
    #--- 変数宣言 ---#

    wb = load_workbook(filename=inputFile, read_only=True)
    ws = wb.active

    # 各行のデータとインデックスをリストにまとめる
    rows_with_index = [[idx] + list(row) for idx, row in enumerate(ws.iter_rows(min_row=minRow, values_only=True), start=startIndex)]
    return rows_with_index


#--- 辞書型リストデータをExcelファイルへ書込 ---#
# data:辞書型リスト
def WriteDicListToExcel(data, outputFile):
    wb = Workbook()
    ws = wb.active

    # ヘッダーを取得
    headers = list(data[0].keys())
    
    # ヘッダーを書き込む
    ws.append(headers)
    
    # データを書き込む
    for item in data:
        row = [item[header] for header in headers]
        ws.append(row)

    # Excelファイルを保存
    wb.save(outputFile)


#--- Excelファイルからキーワードを含む行の抽出 ---#
# このメソッドは将来的に変更修正かけたい
def ExtractDataFromExcel(inputFile, colIndex, keyword):
    #--- 変数宣言 ---#

    # Excelファイルを開く
    wb = load_workbook(inputFile)
    # アクティブなシートを取得する
    ws = wb.active

    # データを保持するリスト
    listData = []

    for row in ws.iter_rows(values_only=True):
        if row[colIndex] == keyword:
            listData.append(row)
    
    return listData


#--- リストデータをExcelファイルへ書込 ---#
# data (list)
def WriteListDataToExcel(data, outputFile, headerStr=''):
    #--- 変数宣言 ---#
    startIndex = 1

    wb = Workbook()
    ws = wb.active

    if (headerStr != ''):
        # ヘッダ行を追加
        for colIndex, header in enumerate(headerStr, start=1):
            ws.cell(row=1, column=colIndex, value=header)
        startIndex = 2
    
    # データリストをExcelファイルに書込
    for idx, row in enumerate(data, start=startIndex): # データ行はstartIndex値から開始
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=idx, column=col_idx, value=value)
            
    # Excelファイルを保存
    wb.save(outputFile)


#=== 設定ファイル読込ブロック ===#
#--- 色情報の設定ファイル(csv)を読み込み、辞書型の変数に値をセットし返却するメソッド ---#
# <Parameters>
# inputFile (str)           : 入力'.csvファイル'(.csvまでのパス)
#
# <Returns>
# colorDict(dict)           : 色情報辞書型変数(色名称:色コード)
def ReadColorInfoFile(inputFile):
    #--- 変数宣言 ---#
    colorDict = {}

    with open(inputFile, 'r', newline='') as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            colorName = row['# ColorName']
            colorCode = row['ColorCode']
            colorDict[colorName] = colorCode

    return colorDict

#--- main関数 ---#
def main():
    print("Youは本スクリプト：{}を直接実行しました".format(curScript))

if __name__ == "__main__":
    main()
