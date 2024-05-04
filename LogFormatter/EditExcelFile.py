#!/usr/local/bin/python3
# coding: utf-8

###
# EidtExcelFile
# エクセルファイル編集スクリプト
# 
###

#--- モジュールのインポート ---#
import sys
from openpyxl import load_workbook, utils, styles


#--- 変数宣言 ---#
curScript = sys.argv[0]
ROW_HEIGHT_RATIO = 1.3


#--- 入力Excelファイル開き、アクティブなシートを取得するメソッド ---#
# inputFile:入力'.xlsxファイル'(.xlsxまでのパス)
def OpenExcelFile(inputFile):
    #--- 変数宣言 ---#

    # Excelファイルを開く
    wb = load_workbook(inputFile)
    # アクティブなシートを取得する
    ws = wb.active

    return wb, ws


#--- セルの列幅の自動調整 ---#
# inputFile:入力'.xlsxファイル'(.xlsxまでのパス)
def AdjustColWidth(inputFile):
    #--- 変数宣言 ---#

    wb, ws = OpenExcelFile(inputFile)

    # セルの幅を自動調整
    for column in ws.columns:
        max_length = 0
        column_letter = utils.get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2 # 適切な調整値を加える
        ws.column_dimensions[column_letter].width = adjusted_width

    # Excelファイルを保存
    wb.save(inputFile)


#--- セルの行高設定(引数指定) ---#
# inputFile:入力'.xlsxファイル'(.xlsxまでのパス)
# hVal:セルの行高
def SetRowHeight(inputFile, hVal):
    #--- 変数宣言 ---#

    wb, ws = OpenExcelFile(inputFile)

    # 行高を設定
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = hVal / ROW_HEIGHT_RATIO # 行高を設定

    # Excelファイルを保存
    wb.save(inputFile)


#--- 全てのセルのフォント設定(引数指定) ---#
# inputFile:入力'.xlsxファイル'(.xlsxまでのパス)
# fontType:フォント名称
def SetDesignedFont(inputFile, fontType):
    #--- 変数宣言 ---#

    wb, ws = OpenExcelFile(inputFile)

    # 全ての行高を設定
    for row in ws.iter_rows():
        for cell in row:
            cell.font = styles.Font(name=fontType) # フォントを設定

    # Excelファイルを保存
    wb.save(inputFile)


#--- Excelファイルにヘッダを挿入する ---#
# inputFile:入力'.xlsxファイル'(.xlsxまでのパス)
# headStrList:ヘッダ文字列リスト
def InsertHeader(inputFile, headStrList):
    #--- 変数宣言 ---#

    wb, ws = OpenExcelFile(inputFile)

    # ヘッダを挿入
    ws.insert_rows(1)
    for col, header in enumerate(headStrList, start=1):
        ws.cell(row=1, column=col, value=header)

    # Excelファイルを保存
    wb.save(inputFile)


#--- Excelファイルの特定の列にフィルタを適用する ---#
# inputFile:入力'.xlsxファイル'(.xlsxまでのパス)
# colS:フィルタ対象スタート列
# colE:フィルタ対象エンド列
def ApplyFilter(inputFile, colS, colE):
    #--- 変数宣言 ---#

    wb, ws = OpenExcelFile(inputFile)

    # フィルタをかける対象の列を指定
    colSLetter = utils.get_column_letter(colS)
    colELetter = utils.get_column_letter(colE)

    # フィルタをかける範囲を指定
    filterRange = f"{colSLetter}:{colELetter}"

    # フィルタを適用
    ws.auto_filter.ref = filterRange

    # Excelファイルを保存
    wb.save(inputFile)


#--- Excelファイルの特定セルでウィンドウを固定する ---#
# inputFile:入力'.xlsxファイル'(.xlsxまでのパス)
# colS:フィルタ対象スタート列
# areaVal:ウィンドウ枠固定セル位置
def UseFreezePanes(inputFile, areaVal):
    #--- 変数宣言 ---#

    wb, ws = OpenExcelFile(inputFile)

    # 引数の値の位置でウィンドウ枠を固定
    ws.freeze_panes = areaVal

    # Excelファイルを保存
    wb.save(inputFile)


#--- Excelファイルのカラムの値に対して行のセルの色を設定するメソッド ---#
# <Parameters>
# inputFile (str)           : 入力'.xlsxファイル'(.xlsxまでのパス)
# colorColumnIndex (int)    : 色情報の列のindex(0から始まる)
# maxColumns (int)          : 変更する列の最大数
# startRow (int, optional)  : 走査開始行のindex(デフォルト値:2)
# endRow (int, optional)    : 走査終了行のindex(デフォルト値:None(最終行まで))
#
# <Returns>
# None
def SetRowColors(inputFile, colorColumnIndex, maxColumns, startRow=2, endRow=None):
    #--- 変数宣言 ---#

    # 色と対応する色コード辞書
    colorCodeDict = {
        "'red'": "F2DCDB", # 赤
        "'blue'": "DDEBF7", # 青
        "'yellow'": "FFF2CC", # 黄
        "'green'": "E2EFDA", # 緑
    }

    wb, ws = OpenExcelFile(inputFile)

    # 各行を走査して、色情報の列を読み取り、該当する色に変更する
    for row in ws.iter_rows(min_row=startRow, max_row=endRow):
        colorInfo = row[colorColumnIndex].value
        if colorInfo in colorCodeDict:
            colorCode = colorCodeDict[colorInfo]
            fill = styles.PatternFill(start_color=colorCode, end_color=colorCode, fill_type="solid")
            for cell in row[:maxColumns]:
                cell.fill = fill

    # Excelファイルを保存
    wb.save(inputFile)

#--- main関数 ---#
def main():
    print("Youは本スクリプト：{}を直接実行しました".format(curScript))

if __name__ == "__main__":
    main()
