#!/usr/local/bin/python3
# coding: utf-8

###
# ExcelFileFormater
# エクセルファイル編集スクリプト
# 
###

#--- モジュールのインポート ---#
import sys
from openpyxl import load_workbook, utils, styles

#--- 変数宣言 ---#
curScript = sys.argv[0]
ROW_HEIGHT_RATIO = 1.3


#--- 入力Excelファイル開き、アクティブなシートを取得するメソッド ---#
# inputFile (str) : 入力'.xlsxファイル'(.xlsxまでのパス)
def GetExcelWbWs(inputFile):
    #--- 変数宣言 ---#

    # Excelファイルを開く
    wb = load_workbook(inputFile)
    # アクティブなシートを取得する
    ws = wb.active

    return wb, ws


#--- セルの列幅の自動調整 ---#
# inputFile (str) : 入力'.xlsxファイル'(.xlsxまでのパス)
def AdjustColWidth(inputFile):
    #--- 変数宣言 ---#

    wb, ws = GetExcelWbWs(inputFile)

    # セルの幅を自動調整
    for column in ws.columns:
        max_length = 0
        column_letter = utils.get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2 # 適切な調整値を加える
        ws.column_dimensions[column_letter].width = adjusted_width

    # Excelファイルを保存
    wb.save(inputFile)


#--- セルの行高設定(引数指定) ---#
# inputFile (str) : 入力'.xlsxファイル'(.xlsxまでのパス)
# hVal :セルの行高
def SetRowHeight(inputFile, hVal):
    #--- 変数宣言 ---#

    wb, ws = GetExcelWbWs(inputFile)

    # 行高を設定
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = hVal / ROW_HEIGHT_RATIO # 行高を設定

    # Excelファイルを保存
    wb.save(inputFile)


#--- 全てのセルのフォント設定(引数指定) ---#
# inputFile (str) : 入力'.xlsxファイル'(.xlsxまでのパス)
# fontType (str) : フォント名称
def SetDesignedFont(inputFile, fontType):
    #--- 変数宣言 ---#

    wb, ws = GetExcelWbWs(inputFile)

    # 全てのセルのフォントを変更する
    for row in ws.iter_rows():
        for cell in row:
            # 現在のフォントの設定を取得する
            currentFont = cell.font
            # フォントを設定する
            cell.font = styles.Font(name=fontType, bold=currentFont.bold, italic=currentFont.italic, color=currentFont.color,
                                    size=currentFont.size, underline=currentFont.underline, strike=currentFont.strike)

    # Excelファイルを保存
    wb.save(inputFile)


#--- Excelファイルにヘッダを挿入する ---#
# inputFile:入力'.xlsxファイル'(.xlsxまでのパス)
# headStrList:ヘッダ文字列リスト
def InsertHeader(inputFile, headStrList):
    #--- 変数宣言 ---#

    wb, ws = GetExcelWbWs(inputFile)

    # ヘッダを挿入
    ws.insert_rows(1)
    for col, header in enumerate(headStrList, start=1):
        ws.cell(row=1, column=col, value=header)

    # Excelファイルを保存
    wb.save(inputFile)


#--- ヘッダ行の文字を太字に設定する ---#
# inputFile:入力'*.xlsxファイル'(*.xlsxまでのパス)
def SetHeaderFontBold(inputFile):
    wb, ws = GetExcelWbWs(inputFile)

    # ヘッダー行のセルのフォントを一時的に保存する
    font_cache = {}
    for cell in ws[1]:
        font_cache[cell.coordinate] = styles.Font(name=cell.font.name, size=cell.font.size)

    # ヘッダー行の文字を太字にする
    for cell in ws[1]:
        cell.font = styles.Font(bold=True, name=font_cache[cell.coordinate].name, size=font_cache[cell.coordinate].size)

    # Excelファイルを保存
    wb.save(inputFile)


#--- ヘッダ行文字列を上下中央揃えに設定する ---#
# inputFile (str) : 入力'*.xlsxファイル'(*.xlsxまでのパス)
def SetHeaderAlignment(inputFile):
    wb, ws = GetExcelWbWs(inputFile)

    # ヘッダ行のセルに対して上下中央揃えを適用
    for cell in ws[1]:  # 1行目がヘッダ行と仮定
        cell.alignment = styles.Alignment(vertical='center', horizontal='center')

    # Excelファイルを保存
    wb.save(inputFile)


#--- 特定のカラム列のセルを右揃えに設定する ---#
# inputFile (str) : 入力'*.xlsxファイル'(*.xlsxまでのパス)
def SetColumnCellsRightAlignment(inputFile, targetCol):
    #--- 変数宣言 ---#
    # 列index値の調整
    targetCol = targetCol + 1

    wb, ws = GetExcelWbWs(inputFile)

    # 特定のカラム列のセルを右揃えに設定
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=targetCol, max_col=targetCol):
        for cell in row:
            cell.alignment = styles.Alignment(vertical='center', horizontal='right')

    # Excelファイルを保存
    wb.save(inputFile)

#--- Excelファイルの特定の列にフィルタを適用する ---#
# inputFile:入力'.xlsxファイル'(.xlsxまでのパス)
# colS:フィルタ対象スタート列
# colE:フィルタ対象エンド列
def ApplyFilter(inputFile, colS, colE):
    #--- 変数宣言 ---#

    wb, ws = GetExcelWbWs(inputFile)

    # フィルタをかける対象の列を指定
    colSLetter = utils.get_column_letter(colS)
    colELetter = utils.get_column_letter(colE)

    # フィルタをかける範囲を指定
    filterRange = f"{colSLetter}:{colELetter}"

    # フィルタを適用
    ws.auto_filter.ref = filterRange

    # Excelファイルを保存
    wb.save(inputFile)


#--- Excelファイルの特定セルでウィンドウを固定する ---#
# inputFile:入力'.xlsxファイル'(.xlsxまでのパス)
# colS:フィルタ対象スタート列
# areaVal:ウィンドウ枠固定セル位置
def UseFreezePanes(inputFile, areaVal):
    #--- 変数宣言 ---#

    wb, ws = GetExcelWbWs(inputFile)

    # 引数の値の位置でウィンドウ枠を固定
    ws.freeze_panes = areaVal

    # Excelファイルを保存
    wb.save(inputFile)


#--- Excelファイルのカラムの値に対して行のセルの色を設定するメソッド ---#
# <Parameters>
# inputFile (str)           : 入力'.xlsxファイル'(.xlsxまでのパス)
# colorCodeDict(dict)       : 色情報辞書型変数(色名称:色コード)
# colorColumnIndex (int)    : 色情報の列のindex(0から始まる)
# maxColumns (int)          : 変更する列の最大数(0から始まる)
# startRow (int, optional)  : 走査開始行のindex(デフォルト値:2)
# endRow (int, optional)    : 走査終了行のindex(デフォルト値:None(最終行まで))
#
# <Returns>
# None
def SetRowColors(inputFile, colorCodeDict ,colorColumnIndex, maxColumns, startRow=2, endRow=None):
    #--- 変数宣言 ---#

    """
    # 色と対応する色コード辞書
    colorCodeDict = {
        "'red'": "F2DCDB", # 赤
        "'blue'": "DDEBF7", # 青
        "'yellow'": "FFF2CC", # 黄
        "'green'": "E2EFDA", # 緑
    }
    """

    # maxColumnsのindex値の調整
    maxColumns = maxColumns + 1

    wb, ws = GetExcelWbWs(inputFile)

    # 各行を走査して、色情報の列を読み取り、該当する色に変更する
    for row in ws.iter_rows(min_row=startRow, max_row=endRow):
        colorInfo = row[colorColumnIndex].value
        if colorInfo in colorCodeDict:
            colorCode = colorCodeDict[colorInfo]
            fill = styles.PatternFill(start_color=colorCode, end_color=colorCode, fill_type="solid")
            for cell in row[:maxColumns]:
                cell.fill = fill

    # Excelファイルを保存
    wb.save(inputFile)

#--- main関数 ---#
def main():
    print("Youは本スクリプト：{}を直接実行しました".format(curScript))

if __name__ == "__main__":
    main()
