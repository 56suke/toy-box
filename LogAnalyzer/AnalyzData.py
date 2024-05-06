#!/usr/local/bin/python3
# coding: utf-8

###
# AnalyzData
# データ分析スクリプト
# 
###

#--- モジュールのインポート ---#
import sys
from openpyxl import load_workbook, Workbook

#--- 自作モジュールのインポート ---#


#--- 変数宣言 ---#
curScript = sys.argv[0]


# DataPairクラスのコンストラクタ
# before及びafterのデータとその行番号を初期化
#
# <Args>
# beforeData (dict) : before時点でのデータ
# afterData (dict)  : after時点でのデータ
# beforeIndex (int) : beforeデータのExcel行番号
# afterIndex (int)  : afterデータのExcel行番号
class DataPair:
    def __init__(self, beforeData, afterData, beforeIndex, afterIndex):
        self.beforeDate = beforeData['Date']
        self.beforeTime = beforeData['Time']
        self.beforeData1 = beforeData['Data1']
        self.beforeData2 = beforeData['Data2']
        self.afterDate = afterData['Date']
        self.afterTime = afterData['Time']
        self.afterData1 = afterData['Data1']
        self.afterData2 = afterData['Data2']
        self.value = beforeData['Value']
        self.beforeIndex = beforeIndex
        self.afterIndex = afterIndex

# ExcelDataProcessorクラス
# コンストラクタ:データペアのリストを初期化し、処理対象のbeforeデータを保持する辞書を初期化
class ExcelDataProcessor:
    def __init__(self):
        self.dataPairs = []
        self.beforeDataDict = {}

    def processExcelFile(self, inputFile):
        #--- 変数宣言 ---#
        wb = load_workbook(inputFile)
        ws = wb.active

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):  # 行番号の追跡
            date, time, data1, data2, _, direction, value = row
            if direction == "'before'":
                if value not in self.beforeDataDict:  # 最初の「before」データのみ保存
                    self.beforeDataDict[value] = {
                        'Date': date,
                        'Time': time,
                        'Data1': data1,
                        'Data2': data2,
                        'Value': value,
                        'Index': idx
                    }
            elif direction == "'after'" and value in self.beforeDataDict:
                before_data = self.beforeDataDict[value]
                after_data = {'Date': date, 'Time': time, 'Data1': data1, 'Data2': data2}
                pair = DataPair(before_data, after_data, before_data['Index'], idx)
                self.dataPairs.append(pair)
                del self.beforeDataDict[value]  # マッチした「before」データを削除

    def WriteDataPairsToExcelFile(self, outputFile):
        #--- 変数宣言 ---#
        wb = Workbook()
        ws = wb.active

        # ヘッダーを書き込む
        header = [
            "# Before Date",
            "Before Time",
            "Before Data1",
            "Before Data2",
            "After Date",
            "After Time",
            "After Data1",
            "After Data2",
            "Value",
            "Before Index",
            "After Index"
            ]
        ws.append(header)

        # データを書き込む
        for pair in self.dataPairs:
            row = [
                pair.beforeDate,
                pair.beforeTime,
                pair.beforeData1,
                pair.beforeData2,
                pair.afterDate,
                pair.afterTime,
                pair.afterData1,
                pair.afterData2,
                pair.value,
                pair.beforeIndex,
                pair.afterIndex
                ]
            ws.append(row)

        # ワークブックを保存する
        wb.save(outputFile)

    def matchBeforeAfterAndWrite(self, inputFile):
        wb = load_workbook(inputFile)
        ws = wb.active

        # ヘッダーを追加するために、既存のヘッダーを拡張
        if ws.cell(row=1, column=9).value is None:
            ws.cell(row=1, column=9).value = 'MatchIndex'

        beforeDataDict = {}
        afterDataCandidates = {}

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if len(row) < 7:
                print(f"Error: Row {idx} has insufficient columns.")
                continue

            direction = row[5]  # 6番目の列が方向
            value = row[6]  # 7番目の列が値

            if direction == "'before'":
                if value not in beforeDataDict:
                    beforeDataDict[value] = idx
                ws.cell(row=idx, column=9).value = '-'

            elif direction == "'after'":
                if value not in afterDataCandidates:
                    afterDataCandidates[value] = []
                afterDataCandidates[value].append(idx)
                ws.cell(row=idx, column=9).value = '-'

        for value, before_index in beforeDataDict.items():
            if value in afterDataCandidates and afterDataCandidates[value]:
                after_index = afterDataCandidates[value].pop(0)
                ws.cell(row=before_index, column=9).value = after_index
                ws.cell(row=after_index, column=9).value = before_index

        for value, indices in afterDataCandidates.items():
            for idx in indices:
                ws.cell(row=idx, column=9).value = '-'

        wb.save("EditFile.xlsx")

#--- main関数 ---#
def main():
    print("Youは本スクリプト：{}を直接実行しました".format(curScript))

    # 使用例
    processor = ExcelDataProcessor()
    processor.processExcelFile("input.xlsx")
    
    """
    for pair in processor.dataPairs:
        print("Before Index:", pair.beforeIndex, "Data:", pair.beforeDate, pair.beforeTime, pair.beforeData1, pair.beforeData2)
        print("After Index:", pair.afterIndex, "Data:", pair.afterDate, pair.afterTime, pair.afterData1, pair.afterData2)
        print("Value:", pair.value)
    """

    processor.WriteDataPairsToExcelFile("output.xlsx")

    processor.matchBeforeAfterAndWrite("input.xlsx")

if __name__ == "__main__":
    main()
